import openpyxl
import os


# Open a workbook from a file or create a new one if it doesn't exist
def open_workbook(file_name):
    if os.path.exists(file_name):
        workbook = openpyxl.load_workbook(file_name)
    else:
        workbook = openpyxl.Workbook()

    # Remove the default sheet if it exists
    default_sheet = workbook.active
    if default_sheet.title == "Sheet":
        workbook.remove(default_sheet)

    return workbook


# Open a worksheet from a workbook or create a new one if it doesn't exist
def open_worksheet(workbook, sheet_name):
    if sheet_name in workbook.sheetnames:
        worksheet = workbook[sheet_name]
    else:
        worksheet = workbook.create_sheet(sheet_name)

    return worksheet


# Save the workbook to a file
def save_workbook(workbook, file_name):
    workbook.save(file_name)


# Write data to a specific cell based on row and column indices
def write_cell(sheet, row, column, data):
    sheet.cell(row=row, column=column).value = data


# Read data from a specific cell based on row and column indices
def read_cell(sheet, row, column):
    return sheet.cell(row=row, column=column).value


# Create a column header in the worksheet or find a place to put it
def create_column_header(sheet, header_text):
    for column in range(1, sheet.max_column + 1):
        if read_cell(sheet, 1, column) == header_text:
            return  # Header already exists, no need to create

    # Header not found, find an empty column to put it
    for column in range(1, sheet.max_column + 2):
        if read_cell(sheet, 1, column) is None:
            write_cell(sheet, row=1, column=column, data=header_text)
            return

    # No empty column found, append the header to the next available column
    next_column = sheet.max_column + 1
    write_cell(sheet, row=1, column=next_column, data=header_text)


# Create a row header in the worksheet or find a place to put it
def create_row_header(sheet, header_text):
    for row in range(1, sheet.max_row + 1):
        if read_cell(sheet, row, 1) == header_text:
            return  # Header already exists, no need to create

    # Header not found, find an empty row to put it
    for row in range(1, sheet.max_row + 2):
        if read_cell(sheet, row, 1) is None:
            write_cell(sheet, row=row, column=1, data=header_text)
            return

    # No empty row found, append the header to the next available row
    next_row = sheet.max_row + 1
    write_cell(sheet, row=next_row, column=1, data=header_text)


# Write data to a specific cell based on row and column headers
def write_cell_by_headers(sheet, row_header, column_header, data):
    row = get_row_index(sheet, row_header)
    column = get_column_index(sheet, column_header)
    write_cell(sheet, row, column, data)


# Read data from a specific cell based on row and column headers
def read_cell_by_headers(sheet, row_header, column_header):
    row = get_row_index(sheet, row_header)
    column = get_column_index(sheet, column_header)
    return read_cell(sheet, row, column)


# Get the row index based on the row header value
def get_row_index(sheet, row_header):
    for row in range(1, sheet.max_row + 1):
        if read_cell(sheet, row, 1) == row_header:
            return row
    raise ValueError("Row header not found.")


# Get the column index based on the column header value
def get_column_index(sheet, column_header):
    for column in range(1, sheet.max_column + 1):
        if read_cell(sheet, 1, column) == column_header:
            return column
    raise ValueError("Column header not found.")


def get_empty_row_index(sheet, column_header):
    # Select the column by index (assuming 'col' is the column index)
    col_index = get_column_index(sheet,
                                 column_header) - 1  # Replace with the desired column index (1 for column A, 2 for column B, etc.)

    column = list(sheet.columns)[col_index]

    # Iterate through the column to find the first index with no value
    first_empty_index = None

    for i, cell in enumerate(column, start=1):
        if cell.value is None:
            first_empty_index = i
            break

    # Check if the first empty index is greater than the first column index
    if first_empty_index is None:
        first_empty_index = -1
    elif first_empty_index > get_last_index(sheet):
        first_empty_index = -1

    return first_empty_index


def get_last_index(sheet):
    i = 0
    for i, cell in enumerate(list(sheet.columns)[0], start=1):
        if cell.value is None:
            return i
    return 1 + i


def has_value(sheet, row_header, column_header):
    return read_cell_by_headers(sheet, row_header, column_header) is None
