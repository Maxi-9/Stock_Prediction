import datetime
import quandl
import pandas as pd
import numpy as np
from datetime import date

import matplotlib.pyplot as plt

import seaborn as sns

from sklearn.metrics import mean_squared_error, r2_score, mean_absolute_error

from sklearn.linear_model import LinearRegression
from sklearn import preprocessing, svm
import yfinance as yf
#replacement for cross validation
from sklearn.model_selection import train_test_split
import openpyxl

import pandas_ta as ta
import Settings

def WriteCell(sheet, row, col, cell_data):
    sheet.cell(row=row,column=col).value = cell_data
    return

def get_current_price(symbol):
    ticker = yf.Ticker(symbol)
    todays_data = ticker.history(period='1d')
    return todays_data['Open'][0]

def LinearPrediction(df,CurrentDate=datetime.date.today(),CurVal=get_current_price(Settings.Stock),testing=False): #CurVal is the morning price 
                                                                                                     # CurrentDate is the current date
    
    #modified version of: https://www.alpharithms.com/predicting-stock-prices-with-linear-regression-214618/
    pd.set_option('mode.chained_assignment', None)
    openVals=df[['Open']].to_dict(orient='index')

    df = df[['Close']]
    #Opendf=df[['Open']]
    df.ta.ema(close='Close', length=10, append=True)

    df = df.iloc[10:]

    X_train, X_test, y_train, y_test = train_test_split(df[['Close']], df[['EMA_10']], test_size=.2)

    model = LinearRegression()
    model.fit(X_train, y_train)
    y_pred = model.predict(X_test)

    print("Model Coefficients(This should be closer to 1):", model.coef_)
    print("Mean Absolute Error(The lower, the better):", mean_absolute_error(y_test, y_pred))
    print("Coefficient of Determination(This should be closer to 1):", r2_score(y_test, y_pred)) 
    
    if testing:
        workbookName = Settings.TestingExcel
    else:
        workbookName = Settings.Excel
        
    try:
        wb =openpyxl.load_workbook(workbookName)
    except:
        wb = openpyxl.Workbook()
        wb.save(workbookName)
    
    workSheetName=f"Linear Regression, {Settings.Stock}"

    try:
        ws = wb[workSheetName]
    except:
        ws = wb.create_sheet(workSheetName)
    #Write Data into Sheet
    WriteCell(ws,1,1,"Dates:")
    WriteCell(ws,1,2,"Open")
    WriteCell(ws,1,3,"Pred Close")
    WriteCell(ws,1,4,"Actual Close")
    WriteCell(ws,1,5,"Buy?")
    WriteCell(ws,1,6,"Profit")

    values=df.to_dict(orient='index')

    cellValue=CurrentDate.strftime("%Y-%m-%d")
    
    row=0
    rowVal=1
    while (True):
        rowVal+=1        
        if ws.cell(row=rowVal,column=1).value==cellValue:
            row=rowVal
            break

        if ws.cell(row=rowVal,column=1).value==None or ws.cell(row=rowVal,column=1).value=="":
            row=rowVal
            break
        
        try:
            timeStamp = datetime.datetime.strptime(ws.cell(row=rowVal,column=1).value,"%Y-%m-%d")
        except:
            print("Error")
            timeStamp=ws.cell(row=rowVal,column=1).value
        
        
        
            
        #WriteCell(ws,rowVal,4,values[timeStamp]["Close"])
        
        
        try:
            timeStamp2 = datetime.datetime.strptime(ws.cell(row=rowVal,column=1).value,"%Y-%m-%d")
        except:
            timeStamp2=ws.cell(row=rowVal,column=1).value
        
        #WriteCell(ws,rowVal,6,values[timeStamp]["Close"]-openVals[timeStamp2]["Open"])
        
        

    print(df.iloc[-1]["EMA_10"])#df['EMA_10'].iloc[[-1]].iloc[0]["EMA_10"])
    EmVal=df.iloc[-1]["EMA_10"]
    
    
    WriteCell(ws,row,1,cellValue)
    WriteCell(ws,row,2,CurVal)
    WriteCell(ws,row,3,""+str(EmVal))
    WriteCell(ws,row,5,""+str(EmVal>CurVal))
    
    wb.save(workbookName)
    
    
